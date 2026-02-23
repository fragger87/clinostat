"""Export simulation results to XLSX and CSV files."""

from __future__ import annotations

import csv
import os

import numpy as np
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.engine import SweepResult


def _write_heatmap_sheet(
    wb: Workbook,
    title: str,
    inner_rpms: np.ndarray,
    outer_rpms: np.ndarray,
    data: np.ndarray,
    number_format: str,
    color_rule: ColorScaleRule,
) -> None:
    """Write a single heatmap sheet: rows = inner RPM, columns = outer RPM."""
    ws = wb.create_sheet(title=title)

    header_font = Font(bold=True)
    center = Alignment(horizontal="center")

    # Top-left corner label
    ws.cell(row=1, column=1, value="Inner \\ Outer").font = header_font
    ws.cell(row=1, column=1).alignment = center

    # Column headers (outer RPMs)
    for j, outer in enumerate(outer_rpms):
        cell = ws.cell(row=1, column=j + 2, value=round(float(outer), 4))
        cell.font = header_font
        cell.alignment = center

    # Row headers (inner RPMs) + data
    for i, inner in enumerate(inner_rpms):
        cell = ws.cell(row=i + 2, column=1, value=round(float(inner), 4))
        cell.font = header_font
        cell.alignment = center

        for j in range(len(outer_rpms)):
            data_cell = ws.cell(row=i + 2, column=j + 2, value=float(data[i, j]))
            data_cell.number_format = number_format
            data_cell.alignment = center

    # Apply conditional color scale to the data range
    data_start = f"B2"
    data_end = f"{get_column_letter(len(outer_rpms) + 1)}{len(inner_rpms) + 1}"
    ws.conditional_formatting.add(f"{data_start}:{data_end}", color_rule)

    # Auto-width columns
    for col_idx in range(1, len(outer_rpms) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


def _write_ranking_sheet(
    wb: Workbook,
    title: str,
    inner_rpms: np.ndarray,
    outer_rpms: np.ndarray,
    magnitudes: np.ndarray,
    distributions: np.ndarray,
) -> None:
    """Write a ranking sheet: all combos sorted by magnitude (ascending)."""
    ws = wb.create_sheet(title=title)

    header_font = Font(bold=True)
    center = Alignment(horizontal="center")

    headers = ["Rank", "Inner RPM", "Outer RPM", "Magnitude (m/s²)", "Distribution Score"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = center

    # Build flat list of all combinations
    rows = []
    for i, inner in enumerate(inner_rpms):
        for j, outer in enumerate(outer_rpms):
            rows.append((float(inner), float(outer), float(magnitudes[i, j]), int(distributions[i, j])))

    # Sort by magnitude ascending (best first)
    rows.sort(key=lambda r: r[2])

    for rank, (inner, outer, mag, dist) in enumerate(rows, 1):
        ws.cell(row=rank + 1, column=1, value=rank).alignment = center
        ws.cell(row=rank + 1, column=2, value=round(inner, 4)).alignment = center
        ws.cell(row=rank + 1, column=3, value=round(outer, 4)).alignment = center
        mag_cell = ws.cell(row=rank + 1, column=4, value=mag)
        mag_cell.number_format = "0.000000"
        mag_cell.alignment = center
        ws.cell(row=rank + 1, column=5, value=dist).alignment = center

    for col_idx in range(1, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def export_sweep_xlsx(result: SweepResult, path: str) -> str:
    """Export sweep results to an XLSX file with heatmap and ranking sheets.

    Creates 3 sheets:
      - Magnitude Heat Map: inner x outer RPM grid of magnitude values (blue=best, red=worst)
      - Distribution Heat Map: inner x outer RPM grid of distribution scores (blue=best, red=worst)
      - Ranking: all combinations sorted by magnitude (ascending)

    Args:
        result: SweepResult from a sweep run.
        path: Output file path (should end in .xlsx).

    Returns:
        Absolute path of the written file.
    """
    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    # Magnitude heat map — lower is better: blue (low) → yellow (mid) → red (high)
    _write_heatmap_sheet(
        wb,
        title="Magnitude Heat Map",
        inner_rpms=result.inner_rpms,
        outer_rpms=result.outer_rpms,
        data=result.magnitudes,
        number_format="0.000000",
        color_rule=ColorScaleRule(
            start_type="min", start_color="4472C4",  # blue (best)
            mid_type="percentile", mid_value=50, mid_color="FFD966",  # yellow
            end_type="max", end_color="C00000",  # red (worst)
        ),
    )

    # Distribution heat map — higher is better: red (low) → yellow (mid) → blue (high)
    _write_heatmap_sheet(
        wb,
        title="Distribution Heat Map",
        inner_rpms=result.inner_rpms,
        outer_rpms=result.outer_rpms,
        data=result.distributions,
        number_format="0",
        color_rule=ColorScaleRule(
            start_type="min", start_color="C00000",  # red (worst)
            mid_type="percentile", mid_value=50, mid_color="FFD966",  # yellow
            end_type="max", end_color="4472C4",  # blue (best)
        ),
    )

    # Ranking sheet
    _write_ranking_sheet(
        wb,
        title="Ranking",
        inner_rpms=result.inner_rpms,
        outer_rpms=result.outer_rpms,
        magnitudes=result.magnitudes,
        distributions=result.distributions,
    )

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    return os.path.abspath(path)


def _write_csv(path: str, header: list[str], rows) -> str:
    """Write rows to a CSV file and return its absolute path."""
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)
    return os.path.abspath(path)


def export_sim_csv(
    t: np.ndarray,
    accel: np.ndarray,
    nongrav: np.ndarray,
    hit_counts: np.ndarray,
    lattice: np.ndarray,
    directory: str,
) -> list[str]:
    """Export single-simulation data to CSV files.

    Creates three files in *directory*:
      - acceleration.csv: time, accel_x, accel_y, accel_z
      - nongrav_acceleration.csv: time, nongrav_x, nongrav_y, nongrav_z
      - hemisphere_hits.csv: lattice_x, lattice_y, lattice_z, hit_count

    Returns:
        List of absolute paths of the written files.
    """
    saved: list[str] = []

    # acceleration.csv
    path = os.path.join(directory, "acceleration.csv")
    rows = ((t[i], accel[i, 0], accel[i, 1], accel[i, 2]) for i in range(len(t)))
    saved.append(_write_csv(path, ["time_s", "accel_x", "accel_y", "accel_z"], rows))

    # nongrav_acceleration.csv
    path = os.path.join(directory, "nongrav_acceleration.csv")
    rows = ((t[i], nongrav[i, 0], nongrav[i, 1], nongrav[i, 2]) for i in range(len(t)))
    saved.append(_write_csv(path, ["time_s", "nongrav_x", "nongrav_y", "nongrav_z"], rows))

    # hemisphere_hits.csv
    path = os.path.join(directory, "hemisphere_hits.csv")
    rows = (
        (lattice[i, 0], lattice[i, 1], lattice[i, 2], int(hit_counts[i]))
        for i in range(len(lattice))
    )
    saved.append(_write_csv(path, ["lattice_x", "lattice_y", "lattice_z", "hit_count"], rows))

    return saved


def export_sweep_csv(result: SweepResult, path: str) -> str:
    """Export sweep results to a CSV file.

    Each row is one RPM combination with columns:
      inner_rpm, outer_rpm, magnitude, distribution

    Args:
        result: SweepResult from a sweep run.
        path: Output file path (should end in .csv).

    Returns:
        Absolute path of the written file.
    """
    rows = (
        (
            float(result.inner_rpms[i]),
            float(result.outer_rpms[j]),
            float(result.magnitudes[i, j]),
            int(result.distributions[i, j]),
        )
        for i in range(len(result.inner_rpms))
        for j in range(len(result.outer_rpms))
    )
    return _write_csv(path, ["inner_rpm", "outer_rpm", "magnitude", "distribution"], rows)
